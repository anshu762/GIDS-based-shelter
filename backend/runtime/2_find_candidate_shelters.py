import math
import time
import json

from collections import defaultdict
from openpyxl import load_workbook


# ==========================================================
# CONFIGURATION
# ==========================================================

EXCEL_FILE = "Dataset1.xlsx"

CANDIDATE_SHEET = "Candidate Shelter"


# ==========================================================
# MAIN FUNCTION
# ==========================================================

def run(scenario_file):

    start_time = time.time()

    print("=" * 60)
    print("FINDING CANDIDATE SHELTERS")
    print("=" * 60)
    print()

    # ------------------------------------------------------
    # LOAD SCENARIO JSON
    # ------------------------------------------------------

    with open(
        scenario_file,
        "r",
        encoding="utf-8"
    ) as file:

        scenario_data = json.load(file)

    scenario = scenario_data["Scenario"]

    epicenter_lat = scenario["Latitude"]
    epicenter_lon = scenario["Longitude"]

    search_radius = scenario["ShelterSearchRadius_km"]

    print(f"Scenario ID            : {scenario['ScenarioID']}")
    print(f"Search Radius          : {search_radius} km")
    print()

    # ------------------------------------------------------
    # LOAD WORKBOOK
    # ------------------------------------------------------

    wb = load_workbook(
        EXCEL_FILE,
        data_only=True
    )

    ws = wb[CANDIDATE_SHEET]

    # ------------------------------------------------------
    # COLUMN INDEX
    # ------------------------------------------------------

    SHELTER_ID_COL = 1
    LOCALITY_ID_COL = 2
    LOCALITY_COL = 3
    OSM_ID_COL = 4
    SHELTER_NAME_COL = 5
    BUILDING_TYPE_COL = 6
    LATITUDE_COL = 7
    LONGITUDE_COL = 8

    # ------------------------------------------------------
    # HAVERSINE
    # ------------------------------------------------------

    def haversine(lat1, lon1, lat2, lon2):

        R = 6371.0

        lat1 = math.radians(lat1)
        lon1 = math.radians(lon1)

        lat2 = math.radians(lat2)
        lon2 = math.radians(lon2)

        dlat = lat2 - lat1
        dlon = lon2 - lon1

        a = (
            math.sin(dlat / 2) ** 2
            +
            math.cos(lat1)
            *
            math.cos(lat2)
            *
            math.sin(dlon / 2) ** 2
        )

        c = 2 * math.atan2(
            math.sqrt(a),
            math.sqrt(1 - a)
        )

        return R * c

    # ------------------------------------------------------
    # VARIABLES
    # ------------------------------------------------------

    candidate_shelters = []

    building_summary = defaultdict(int)

    shelters_checked = 0

    nearest_shelter = None
    farthest_shelter = None

    min_distance = float("inf")
    max_distance = 0

        # ------------------------------------------------------
    # PROCESS ALL SHELTERS
    # ------------------------------------------------------

    for row in range(2, ws.max_row + 1):

        shelter_id = ws.cell(row, SHELTER_ID_COL).value
        locality_id = ws.cell(row, LOCALITY_ID_COL).value
        locality = ws.cell(row, LOCALITY_COL).value
        osm_id = ws.cell(row, OSM_ID_COL).value
        shelter_name = ws.cell(row, SHELTER_NAME_COL).value
        building_type = ws.cell(row, BUILDING_TYPE_COL).value
        latitude = ws.cell(row, LATITUDE_COL).value
        longitude = ws.cell(row, LONGITUDE_COL).value

        if (
            shelter_id is None
            or latitude is None
            or longitude is None
        ):
            continue

        shelters_checked += 1

        distance = haversine(
            epicenter_lat,
            epicenter_lon,
            float(latitude),
            float(longitude)
        )

        if distance <= search_radius:

            shelter = {

                "ShelterID": shelter_id,
                "LocalityID": locality_id,
                "Locality": locality,
                "OSM_ID": osm_id,
                "ShelterName": shelter_name,
                "BuildingType": building_type,
                "Latitude": latitude,
                "Longitude": longitude,
                "Distance(km)": round(distance, 3)

            }

            candidate_shelters.append(shelter)

            building_summary[building_type] += 1

            if distance < min_distance:

                min_distance = distance
                nearest_shelter = shelter

            if distance > max_distance:

                max_distance = distance
                farthest_shelter = shelter

    # ------------------------------------------------------
    # SORT BY DISTANCE
    # ------------------------------------------------------

    candidate_shelters.sort(
        key=lambda x: x["Distance(km)"]
    )

    # ------------------------------------------------------
    # MODULE RESULT
    # ------------------------------------------------------

    module2_result = {

        "candidate_shelters_found": len(candidate_shelters),

        "candidate_shelters": candidate_shelters,

        "building_type_summary": dict(building_summary),

        "execution_time": round(
            time.time() - start_time,
            3
        )

    }

    # ------------------------------------------------------
    # APPEND TO JSON
    # ------------------------------------------------------

    scenario_data["Modules"]["Module2"] = module2_result

    with open(

        scenario_file,

        "w",

        encoding="utf-8"

    ) as file:

        json.dump(

            scenario_data,

            file,

            indent=4

        )

    # ------------------------------------------------------
    # PRINT REPORT
    # ------------------------------------------------------

    print("-" * 60)

    print(f"Shelters Checked        : {shelters_checked:,}")
    print(f"Candidate Shelters      : {len(candidate_shelters):,}")
    print(f"Execution Time          : {module2_result['execution_time']} sec")

    print()

    print("-" * 60)
    print("Building Type Summary")
    print("-" * 60)

    print()

    for building_type, count in sorted(

        building_summary.items(),

        key=lambda x: x[1],

        reverse=True

    ):

        print(f"{building_type:<30}{count:>10}")

    print()

    if nearest_shelter:

        print("-" * 60)
        print("Nearest Shelter")
        print("-" * 60)

        print()

        print(f"Name      : {nearest_shelter['ShelterName']}")
        print(f"Distance  : {nearest_shelter['Distance(km)']} km")

        print()

        print("-" * 60)
        print("Farthest Shelter")
        print("-" * 60)

        print()

        print(f"Name      : {farthest_shelter['ShelterName']}")
        print(f"Distance  : {farthest_shelter['Distance(km)']} km")

    print()

    print("=" * 60)
    print("MODULE 2 COMPLETED SUCCESSFULLY")
    print("=" * 60)

    return scenario_file


# ==========================================================
# STANDALONE EXECUTION
# ==========================================================

if __name__ == "__main__":

    #TEST_JSON = "scenarios/REPLACE_WITH_YOUR_SCENARIO.json"
    TEST_JSON = "scenarios/Flood_Dharavi_5km_20260812_174629.json"
    run(TEST_JSON)