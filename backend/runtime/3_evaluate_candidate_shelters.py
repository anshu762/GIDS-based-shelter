import time
import json

from collections import defaultdict
from openpyxl import load_workbook


# ==========================================================
# CONFIGURATION
# ==========================================================

EXCEL_FILE = "Dataset1.xlsx"

BUILDING_TYPE_MASTER_SHEET = "Building Type Master"


# ==========================================================
# MAIN FUNCTION
# ==========================================================

def run(scenario_file):

    start_time = time.time()

    print("=" * 60)
    print("EVALUATING CANDIDATE SHELTERS")
    print("=" * 60)
    print()

    # ======================================================
    # LOAD SCENARIO JSON
    # ======================================================

    with open(
        scenario_file,
        "r",
        encoding="utf-8"
    ) as file:

        scenario_data = json.load(file)

    scenario = scenario_data["Scenario"]

    disaster_type = scenario["DisasterType"]

    print(f"Scenario ID            : {scenario['ScenarioID']}")
    print(f"Disaster Type           : {disaster_type}")
    print()

    # ======================================================
    # LOAD MODULE 2 RESULTS
    # ======================================================

    if "Module2" not in scenario_data["Modules"]:

        raise ValueError(
            "Module2 results not found in scenario JSON."
        )

    candidate_shelters = scenario_data[
        "Modules"
    ][
        "Module2"
    ][
        "candidate_shelters"
    ]

    print(
        f"Candidate Shelters Received : "
        f"{len(candidate_shelters):,}"
    )

    print()

    # ======================================================
    # LOAD WORKBOOK
    # ======================================================

    print("Opening Workbook...")

    wb = load_workbook(
        EXCEL_FILE,
        data_only=True
    )

    ws_master = wb[
        BUILDING_TYPE_MASTER_SHEET
    ]

    print("Workbook Loaded.")
    print()

    # ======================================================
    # BUILD BUILDING TYPE LOOKUP
    # ======================================================

    building_type_lookup = {}

    for row in range(
        2,
        ws_master.max_row + 1
    ):

        building_type = ws_master.cell(
            row,
            1
        ).value

        if building_type is None:
            continue

        building_type_lookup[
            str(building_type).strip().lower()
        ] = {

            "BuildingType": building_type,

            "DefaultCapacity":
                ws_master.cell(row, 2).value,

            "FloodSafe":
                ws_master.cell(row, 3).value,

            "EarthquakeSafe":
                ws_master.cell(row, 4).value,

            "FireSafe":
                ws_master.cell(row, 5).value,

            "CycloneSafe":
                ws_master.cell(row, 6).value,

            "TemporaryShelter":
                ws_master.cell(row, 7).value,

            "MedicalFacility":
                ws_master.cell(row, 8).value,

            "IndoorOutdoor":
                ws_master.cell(row, 9).value,

            "PriorityWeight":
                ws_master.cell(row, 10).value,

            "StructuralPriority":
                ws_master.cell(row, 11).value,

            "PreferredDisaster":
                ws_master.cell(row, 12).value,

            "Remarks":
                ws_master.cell(row, 13).value
        }

    # ======================================================
    # DISASTER → SAFETY COLUMN MAPPING
    # ======================================================

    disaster_safety_column = {

        "Flood": "FloodSafe",

        "Earthquake": "EarthquakeSafe",

        "Fire": "FireSafe",

        "Cyclone": "CycloneSafe"

    }

    if disaster_type not in disaster_safety_column:

        raise ValueError(
            f"Unsupported disaster type: "
            f"{disaster_type}"
        )

    safety_column = disaster_safety_column[
        disaster_type
    ]

    print(
        f"Safety Rule Used       : "
        f"{safety_column}"
    )

    print()

    # ======================================================
    # VARIABLES
    # ======================================================

    suitable_shelters = []

    unsuitable_shelters = []

    rejection_summary = defaultdict(int)

    building_summary = defaultdict(int)

    shelters_checked = 0

    # ======================================================
    # EVALUATE CANDIDATE SHELTERS
    # ======================================================

    for shelter in candidate_shelters:

        shelters_checked += 1

        building_type = shelter.get(
            "BuildingType"
        )

        # --------------------------------------------------
        # CHECK BUILDING TYPE EXISTS
        # --------------------------------------------------

        if building_type is None:

            reason = (
                "Building type is missing"
            )

            unsuitable_shelter = dict(
                shelter
            )

            unsuitable_shelter[
                "Status"
            ] = "Unsuitable"

            unsuitable_shelter[
                "RejectionReason"
            ] = reason

            unsuitable_shelters.append(
                unsuitable_shelter
            )

            rejection_summary[
                reason
            ] += 1

            continue

        building_key = str(
            building_type
        ).strip().lower()

        building_info = building_type_lookup.get(
            building_key
        )

        # --------------------------------------------------
        # CHECK BUILDING TYPE MASTER
        # --------------------------------------------------

        if building_info is None:

            reason = (
                "Building type not found "
                "in Building Type Master"
            )

            unsuitable_shelter = dict(
                shelter
            )

            unsuitable_shelter[
                "Status"
            ] = "Unsuitable"

            unsuitable_shelter[
                "RejectionReason"
            ] = reason

            unsuitable_shelters.append(
                unsuitable_shelter
            )

            rejection_summary[
                reason
            ] += 1

            continue

        # --------------------------------------------------
        # CHECK DISASTER SAFETY
        # --------------------------------------------------

        safety_status = building_info[
            safety_column
        ]

        if str(
            safety_status
        ).strip().lower() != "yes":

            reason = (
                f"Not marked safe for "
                f"{disaster_type}"
            )

            unsuitable_shelter = dict(
                shelter
            )

            unsuitable_shelter[
                "Status"
            ] = "Unsuitable"

            unsuitable_shelter[
                "RejectionReason"
            ] = reason

            unsuitable_shelters.append(
                unsuitable_shelter
            )

            rejection_summary[
                reason
            ] += 1

            continue

        # --------------------------------------------------
        # SUITABLE SHELTER
        # --------------------------------------------------

        suitable_shelter = dict(
            shelter
        )

        suitable_shelter[
            "Status"
        ] = "Suitable"

        # Building information
        suitable_shelter[
            "DefaultCapacity"
        ] = building_info[
            "DefaultCapacity"
        ]

        suitable_shelter[
            "TemporaryShelter"
        ] = building_info[
            "TemporaryShelter"
        ]

        suitable_shelter[
            "MedicalFacility"
        ] = building_info[
            "MedicalFacility"
        ]

        suitable_shelter[
            "IndoorOutdoor"
        ] = building_info[
            "IndoorOutdoor"
        ]

        suitable_shelter[
            "PriorityWeight"
        ] = building_info[
            "PriorityWeight"
        ]

        suitable_shelter[
            "StructuralPriority"
        ] = building_info[
            "StructuralPriority"
        ]

        suitable_shelter[
            "PreferredDisaster"
        ] = building_info[
            "PreferredDisaster"
        ]

        suitable_shelter[
            "Remarks"
        ] = building_info[
            "Remarks"
        ]

        suitable_shelters.append(
            suitable_shelter
        )

        building_summary[
            building_type
        ] += 1

    # ======================================================
    # SORT SUITABLE SHELTERS
    # ======================================================

    suitable_shelters.sort(

        key=lambda x: (

            -x["PriorityWeight"],

            -x["StructuralPriority"],

            x["Distance(km)"]

        )

    )

    # ======================================================
    # MODULE 3 RESULT
    # ======================================================

    execution_time = round(
        time.time() - start_time,
        3
    )

    module3_result = {

        "candidate_shelters_checked":
            shelters_checked,

        "suitable_shelter_count":
            len(suitable_shelters),

        "unsuitable_shelter_count":
            len(unsuitable_shelters),

        "suitable_shelters":
            suitable_shelters,

        "unsuitable_shelters":
            unsuitable_shelters,

        "building_type_summary":
            dict(building_summary),

        "rejection_summary":
            dict(rejection_summary),

        "execution_time":
            execution_time
    }

    # ======================================================
    # APPEND MODULE 3 TO SCENARIO JSON
    # ======================================================

    scenario_data[
        "Modules"
    ][
        "Module3"
    ] = module3_result

    with open(

        scenario_file,

        "w",

        encoding="utf-8"

    ) as file:

        json.dump(

            scenario_data,

            file,

            indent=4

        )

    # ======================================================
    # PRINT REPORT
    # ======================================================

    print("=" * 60)
    print("MODULE 3 RESULTS")
    print("=" * 60)

    print()

    print(
        f"Candidate Shelters Checked : "
        f"{shelters_checked:,}"
    )

    print(
        f"Suitable Shelters           : "
        f"{len(suitable_shelters):,}"
    )

    print(
        f"Unsuitable Shelters         : "
        f"{len(unsuitable_shelters):,}"
    )

    print(
        f"Execution Time              : "
        f"{execution_time} sec"
    )

    print()

    # ======================================================
    # BUILDING TYPE SUMMARY
    # ======================================================

    print("-" * 60)
    print("Suitable Shelter Summary")
    print("-" * 60)

    print()

    print(
        f"{'Building Type':<25}"
        f"{'Count':>10}"
    )

    print("-" * 40)

    for building_type, count in sorted(

        building_summary.items(),

        key=lambda x: x[1],

        reverse=True

    ):

        print(

            f"{building_type:<25}"
            f"{count:>10,}"

        )

    print()

    # ======================================================
    # REJECTION SUMMARY
    # ======================================================

    print("-" * 60)
    print("Rejection Summary")
    print("-" * 60)

    print()

    if rejection_summary:

        for reason, count in sorted(

            rejection_summary.items(),

            key=lambda x: x[1],

            reverse=True

        ):

            print(
                f"{reason:<45}"
                f"{count:>10,}"
            )

    else:

        print("No shelters rejected.")

    print()

    # ======================================================
    # TOP CANDIDATES
    # ======================================================

    print("-" * 60)
    print("Top Suitable Candidates")
    print("-" * 60)

    print()

    for shelter in suitable_shelters[:10]:

        print(

            f"{shelter['ShelterID']} | "

            f"{shelter['ShelterName']} | "

            f"{shelter['BuildingType']} | "

            f"Capacity: "
            f"{shelter['DefaultCapacity']} | "

            f"Priority: "
            f"{shelter['PriorityWeight']} | "

            f"Distance: "
            f"{shelter['Distance(km)']} km"

        )

    print()

    print("=" * 60)
    print("MODULE 3 COMPLETED SUCCESSFULLY")
    print("=" * 60)

    return scenario_file


# ==========================================================
# STANDALONE EXECUTION
# ==========================================================

if __name__ == "__main__":

    # TEST_JSON = "scenarios/REPLACE_WITH_YOUR_SCENARIO.json"

    TEST_JSON = (
        "scenarios/Flood_Dharavi_5km_20260812_174629.json"
    )

    run(TEST_JSON)