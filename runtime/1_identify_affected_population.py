import math
import time
from collections import defaultdict
from openpyxl import load_workbook
import json
import os
from datetime import datetime
def run(disaster_type, epicenter_name,epicenter_lat, epicenter_lon, radius_km):

    # ==========================================================
    # CONFIGURATION
    # ==========================================================

    EXCEL_FILE = "Dataset1.xlsx"

    LOCALITY_SHEET = "Localities"
    POPULATION_NODE_SHEET = "Population Nodes"

    # ==========================================================
    # USER INPUT (Temporary)
    # ==========================================================

    DISASTER_TYPE = disaster_type

    EPICENTER_NAME = epicenter_name

    EPICENTER_LAT = epicenter_lat
    EPICENTER_LON = epicenter_lon

    RADIUS_KM =  radius_km
    
    # ==========================================================
    # LOAD WORKBOOK
    # ==========================================================

    print("Opening Workbook...")

    wb = load_workbook(EXCEL_FILE, data_only=True)

    ws_locality = wb[LOCALITY_SHEET]
    ws_nodes = wb[POPULATION_NODE_SHEET]

    print("Workbook Loaded.\n")

    # ==========================================================
    # COLUMN INDEX
    # ==========================================================

    NODE_ID_COL = 1
    LOCALITY_ID_COL = 2
    LATITUDE_COL = 3
    LONGITUDE_COL = 4
    POPULATION_COL = 5

    LOCALITY_ID_MASTER_COL = 4
    LOCALITY_NAME_MASTER_COL = 5

    # ==========================================================
    # BUILD LOCALITY LOOKUP
    # ==========================================================

    locality_lookup = {}

    for row in range(2, ws_locality.max_row + 1):

        locality_id = ws_locality.cell(row, LOCALITY_ID_MASTER_COL).value
        locality_name = ws_locality.cell(row, LOCALITY_NAME_MASTER_COL).value

        locality_lookup[locality_id] = locality_name

    # ==========================================================
    # HAVERSINE DISTANCE
    # ==========================================================

    def haversine(lat1, lon1, lat2, lon2):

        R = 6371.0

        lat1 = math.radians(lat1)
        lon1 = math.radians(lon1)

        lat2 = math.radians(lat2)
        lon2 = math.radians(lon2)

        dlat = lat2 - lat1
        dlon = lon2 - lon1

        a = (
            math.sin(dlat / 2) ** 2
            +
            math.cos(lat1)
            *
            math.cos(lat2)
            *
            math.sin(dlon / 2) ** 2
        )

        c = 2 * math.atan2(
            math.sqrt(a),
            math.sqrt(1 - a)
        )

        return R * c

    # ==========================================================
    # START TIMER
    # ==========================================================

    start_time = time.time()

    affected_nodes = []

    affected_localities = set()

    locality_summary = defaultdict(

        lambda: {

            "Nodes": 0,

            "Population": 0

        }

    )

    total_population = 0

    nodes_checked = 0

    nearest_node = None

    farthest_node = None

    min_distance = float("inf")

    max_distance = 0

    print("=" * 60)
    print("IDENTIFYING AFFECTED POPULATION")
    print("=" * 60)
    print()

    # ==========================================================
    # PROCESS POPULATION NODES
    # ==========================================================

    for row in range(2, ws_nodes.max_row + 1):

        node_id = ws_nodes.cell(row, NODE_ID_COL).value

        locality_id = ws_nodes.cell(row, LOCALITY_ID_COL).value

        latitude = ws_nodes.cell(row, LATITUDE_COL).value

        longitude = ws_nodes.cell(row, LONGITUDE_COL).value

        population = ws_nodes.cell(row, POPULATION_COL).value

        if (

            node_id is None or

            locality_id is None or

            latitude is None or

            longitude is None or

            population is None

        ):

            continue

        nodes_checked += 1

        distance = haversine(

            EPICENTER_LAT,

            EPICENTER_LON,

            float(latitude),

            float(longitude)

        )

        if distance <= RADIUS_KM:

            locality_name = locality_lookup.get(

                locality_id,

                "Unknown"

            )

            node = {

                "PopulationNodeID": node_id,

                "LocalityID": locality_id,

                "Locality": locality_name,

                "Latitude": latitude,

                "Longitude": longitude,

                "Population": population,

                "Distance(km)": round(distance, 3)

            }

            affected_nodes.append(node)

            affected_localities.add(locality_name)

            total_population += population

            locality_summary[locality_name]["Nodes"] += 1

            locality_summary[locality_name]["Population"] += population

            if distance < min_distance:

                min_distance = distance

                nearest_node = node

            if distance > max_distance:

                max_distance = distance

                farthest_node = node
    # ==========================================================
    # SORT AFFECTED NODES
    # ==========================================================

    affected_nodes.sort(

        key=lambda x: x["Distance(km)"]

    )

    # ==========================================================
    # STOP TIMER
    # ==========================================================

    execution_time = round(

        time.time() - start_time,

        3

    )

    # ==========================================================
    # RESULT OBJECT
    # ==========================================================

    result = {

        "affected_population": total_population,

        "affected_nodes": affected_nodes,

        "affected_localities": sorted(
            list(affected_localities)
        ),

        "execution_time": execution_time

    }
    # ==========================================================
    # CREATE SCENARIO DIRECTORY
    # ==========================================================

    SCENARIO_FOLDER = "scenarios"

    os.makedirs(SCENARIO_FOLDER, exist_ok=True)

    # ==========================================================
    # GENERATE SCENARIO FILE NAME
    # ==========================================================

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    SCENARIO_ID = f"{DISASTER_TYPE}_{EPICENTER_NAME}_{int(RADIUS_KM)}km_{timestamp}"

    scenario_file = os.path.join(

        SCENARIO_FOLDER,

        f"{SCENARIO_ID}.json"


    )

    # ==========================================================
    # CREATE JSON STRUCTURE
    # ==========================================================

    scenario_data = {
    "Scenario": {

        "ScenarioID": SCENARIO_ID,

        "DisasterType": DISASTER_TYPE,

        "Epicenter": EPICENTER_NAME,

        "Latitude": EPICENTER_LAT,

        "Longitude": EPICENTER_LON,

        "DisasterRadius_km": RADIUS_KM,

        "ShelterSearchRadius_km": round(RADIUS_KM * 1.4, 2)

    },
        "Modules":{

        "Module1": result
        }

    }

    # ==========================================================
    # SAVE JSON
    # ==========================================================

    with open(

        scenario_file,

        "w",

        encoding="utf-8"

    ) as file:

        json.dump(

            scenario_data,

            file,

            indent=4

        )

    print()

    print(f"Scenario JSON Created : {scenario_file}")

    # ==========================================================
    # PRINT REPORT
    # ==========================================================

    print("=" * 60)
    print("AFFECTED POPULATION REPORT")
    print("=" * 60)
    print()

    print(f"Disaster Type          : {DISASTER_TYPE}")
    print(f"Epicenter              : {EPICENTER_NAME}")
    print(f"Latitude               : {EPICENTER_LAT:.6f}")
    print(f"Longitude              : {EPICENTER_LON:.6f}")
    print(f"Affected Radius        : {RADIUS_KM:.1f} km")

    print()

    print("-" * 60)

    print(f"Population Nodes Checked : {nodes_checked:,}")
    print(f"Affected Nodes          : {len(affected_nodes):,}")
    print(f"Affected Population     : {total_population:,}")
    print(f"Affected Localities     : {len(affected_localities)}")
    print(f"Execution Time          : {execution_time} sec")

    print()

    print("-" * 60)
    print("Affected Locality Summary")
    print("-" * 60)

    print()

    print(
        f"{'Locality':<25}"
        f"{'Nodes':>10}"
        f"{'Population':>15}"
    )

    print("-" * 50)

    # ==========================================================
    # SORT LOCALITIES BY POPULATION
    # ==========================================================

    sorted_localities = sorted(

        locality_summary.items(),

        key=lambda x: x[1]["Population"],

        reverse=True

    )

    for locality, data in sorted_localities:

        print(

            f"{locality:<25}"

            f"{data['Nodes']:>10,}"

            f"{data['Population']:>15,}"

        )

    print()

    print("-" * 60)
    print("Statistics")
    print("-" * 60)
    print()

    if len(affected_nodes) > 0:

        avg_population = round(

            total_population / len(affected_nodes),

            2

        )

        print(
            f"Average Population / Node : {avg_population}"
        )

        print()

        print(

            f"Nearest Node              : "

            f"{nearest_node['PopulationNodeID']} "

            f"({nearest_node['Locality']}, "

            f"{nearest_node['Distance(km)']} km)"

        )

        print(

            f"Farthest Node             : "

            f"{farthest_node['PopulationNodeID']} "

            f"({farthest_node['Locality']}, "

            f"{farthest_node['Distance(km)']} km)"

        )

    else:

        print("No affected population nodes found.")

    print()

    print("=" * 60)
    print("MODULE 1 COMPLETED SUCCESSFULLY")
    print("=" * 60)
    return scenario_file

if __name__ == "__main__":
    run()